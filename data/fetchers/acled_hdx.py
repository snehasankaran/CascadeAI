"""ACLED conflict data via HDX (no authentication required).

ACLED publishes per-country political-violence rollups on the Humanitarian
Data Exchange (https://data.humdata.org) — the data is the **same dataset**
ACLED's authenticated API serves, only delivered as a weekly-refreshed XLSX
file rather than a live JSON endpoint. HDX downloads require no API key.

This module is the second of three transports for CascadeAI's Action
Verifier:

  1. ACLED v3 API           — when ``ACLED_API_KEY`` + ``ACLED_EMAIL`` are
                              set. Real-time, event-level.
  2. **ACLED via HDX (this)** — when no API credentials are present. Weekly
                              refresh, monthly aggregates, but still live
                              ACLED data with the licensed source on every
                              record.
  3. Static priors          — last-resort fallback if both above fail.

The XLSX schema (validated 2026-05-13 against May-12-2026 release):

    Sheet "Data": Country | Month | Year | Events | Fatalities

We download the **political_violence_events_and_fatalities_by_month-year**
resource for each country, cache the bytes on disk for 7 days (HDX refreshes
weekly anyway), and compute 30-day / 90-day / trend aggregates the Verifier
can reason over.

Per-country file sizes range from ~15 KB (Kenya) to ~3 MB (Sudan). Total
cold-start download for the East Africa region is roughly 6 MB.
"""

from __future__ import annotations

import calendar
import datetime as _dt
import io
import os
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import httpx

# ---------------------------------------------------------------------------
# HDX package discovery
# ---------------------------------------------------------------------------

HDX_BASE = "https://data.humdata.org/api/3/action"

# country (lower-case) → HDX package id. Pattern is
# ``<country-name-with-hyphens>-acled-conflict-data`` but we hard-code the
# 14 countries in CascadeAI's roster so we don't need a search round-trip
# on every Verifier run.
COUNTRY_TO_PACKAGE: dict[str, str] = {
    "kenya": "kenya-acled-conflict-data",
    "ethiopia": "ethiopia-acled-conflict-data",
    "somalia": "somalia-acled-conflict-data",
    "sudan": "sudan-acled-conflict-data",
    "south sudan": "south-sudan-acled-conflict-data",
    "egypt": "egypt-acled-conflict-data",
    "bangladesh": "bangladesh-acled-conflict-data",
    "india": "india-acled-conflict-data",
    "dr congo": "democratic-republic-of-congo-acled-conflict-data",
    "democratic republic of the congo": "democratic-republic-of-congo-acled-conflict-data",
    "drc": "democratic-republic-of-congo-acled-conflict-data",
    "chile": "chile-acled-conflict-data",
    "indonesia": "indonesia-acled-conflict-data",
    "pakistan": "pakistan-acled-conflict-data",
    "yemen": "yemen-acled-conflict-data",
    "myanmar": "myanmar-acled-conflict-data",
    "burma": "myanmar-acled-conflict-data",
}

# Loose region → country list for region-level Verifier queries.
REGION_TO_COUNTRIES: dict[str, list[str]] = {
    "east africa": ["kenya", "ethiopia", "somalia", "sudan", "south sudan"],
    "horn of africa": ["ethiopia", "somalia", "sudan", "south sudan"],
    "north africa": ["egypt", "sudan"],
    "middle east": ["yemen"],
    "south asia": ["bangladesh", "india", "pakistan", "myanmar"],
    "southeast asia": ["myanmar", "indonesia"],
    "latin america": ["chile"],
}

# ---------------------------------------------------------------------------
# HTTP config
# ---------------------------------------------------------------------------

PROXY = os.getenv("HTTPS_PROXY", os.getenv("HTTP_PROXY", "")) or None

# Corporate proxies sometimes inject their own root cert that openssl
# can't validate. When CASCADEAI_VERIFY_SSL=false, we skip cert verify on
# the AWS S3 redirect that HDX downloads land on. Default: verify on.
VERIFY_SSL = os.getenv("CASCADEAI_VERIFY_SSL", "true").strip().lower() != "false"

USER_AGENT = "CascadeAI/1.0 (humanitarian crisis cascade modelling)"

# 7-day on-disk cache — HDX refreshes ACLED data weekly anyway, so a 7d TTL
# means at most one download per country per week of dashboard activity.
CACHE_TTL_SECONDS = 7 * 24 * 3600

CACHE_DIR = Path(__file__).resolve().parents[1] / "cache" / "acled_hdx"
CACHE_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Data shapes
# ---------------------------------------------------------------------------

@dataclass
class _MonthlyRow:
    country: str
    year: int
    month: int  # 1..12
    events: int
    fatalities: int

    @property
    def ordinal(self) -> int:
        return self.year * 12 + (self.month - 1)


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------

def fetch_country_conflict_summary(country: str, days: int = 30) -> Optional[dict]:
    """Return an ACLED-shaped summary for ``country`` from HDX.

    Returns ``None`` if the country isn't in CascadeAI's HDX roster or
    every fetch path fails — the caller (``acled_api.py``) is responsible
    for falling back to static priors.
    """
    pkg = COUNTRY_TO_PACKAGE.get(country.strip().lower())
    if not pkg:
        return None

    rows = _load_country_rows(country, pkg)
    if not rows:
        return None

    # ACLED publishes the in-progress month with partial data (e.g. May 2026
    # mid-month shows ~1/2 of its eventual count). To keep the ``trend`` label
    # honest we exclude any month-in-progress from the trend baseline.
    today = _dt.date.today()
    latest = rows[-1]
    month_in_progress = latest.year == today.year and latest.month == today.month
    complete_rows = rows[:-1] if month_in_progress else rows

    months_window = max(1, days // 30)
    last_n = _aggregate_last_n_months(rows, months_window)
    annual = _aggregate_last_n_months(rows, 12)

    # Trend: last 3 complete months vs the 3 before them.
    recent3 = _aggregate_last_n_months(complete_rows, 3)
    prior3 = _aggregate_last_n_months(complete_rows, 3, skip_first_n=3)
    recent_pm = recent3["events"] / 3 if recent3["events"] else 0
    prior_pm = prior3["events"] / 3 if prior3["events"] else 0
    if prior_pm <= 0:
        trend = "baseline_unknown"
    elif recent_pm > 1.3 * prior_pm:
        trend = "escalating"
    elif recent_pm < 0.7 * prior_pm:
        trend = "de-escalating"
    else:
        trend = "stable"

    label_month = calendar.month_name[latest.month]
    return {
        "source": "ACLED via HDX",
        "country": country.title(),
        "events_30d": last_n["events"],
        "fatalities_30d": last_n["fatalities"],
        "events_90d": _aggregate_last_n_months(rows, 3)["events"],
        "fatalities_90d": _aggregate_last_n_months(rows, 3)["fatalities"],
        "events_annual": annual["events"],
        "fatalities_annual": annual["fatalities"],
        "latest_month_label": f"{label_month} {latest.year}"
            + (" (partial)" if month_in_progress else ""),
        "latest_month_events": latest.events,
        "latest_month_fatalities": latest.fatalities,
        "latest_month_partial": month_in_progress,
        "trend": trend,
        "active_conflicts": [
            f"{country.title()} — political violence ({latest.events} events, "
            f"{latest.fatalities} fatalities in {label_month} {latest.year}"
            + (" — partial" if month_in_progress else "")
            + f"; trend: {trend})"
        ],
        "license": "ACLED · CC BY 4.0 (via HDX)",
        "dataset_url": f"https://data.humdata.org/dataset/{pkg}",
    }


def fetch_region_conflict_summary(region: str, days: int = 30) -> Optional[dict]:
    """Aggregate per-country HDX ACLED rollups across a region.

    Returns ``None`` if the region isn't recognised — the caller falls back
    to static priors.
    """
    countries = REGION_TO_COUNTRIES.get(region.strip().lower())
    if not countries:
        return None

    sub_summaries: list[dict] = []
    for c in countries:
        summary = fetch_country_conflict_summary(c, days=days)
        if summary:
            sub_summaries.append(summary)

    if not sub_summaries:
        return None

    events_30d = sum(s["events_30d"] for s in sub_summaries)
    fatalities_30d = sum(s["fatalities_30d"] for s in sub_summaries)
    events_90d = sum(s["events_90d"] for s in sub_summaries)
    fatalities_90d = sum(s["fatalities_90d"] for s in sub_summaries)

    return {
        "source": "ACLED via HDX",
        "region": region,
        "country_coverage": [s["country"] for s in sub_summaries],
        "events_30d": events_30d,
        "fatalities_30d": fatalities_30d,
        "events_90d": events_90d,
        "fatalities_90d": fatalities_90d,
        "active_conflicts": [c for s in sub_summaries for c in s["active_conflicts"]],
        "license": "ACLED · CC BY 4.0 (via HDX)",
    }


# ---------------------------------------------------------------------------
# Internals — package resolution, download, parse, cache
# ---------------------------------------------------------------------------


def _load_country_rows(country: str, package_id: str) -> list[_MonthlyRow]:
    cache_path = CACHE_DIR / f"{package_id}.xlsx"
    if cache_path.exists() and (time.time() - cache_path.stat().st_mtime) < CACHE_TTL_SECONDS:
        try:
            return _parse_xlsx(country, cache_path.read_bytes())
        except Exception:  # noqa: BLE001 — cache miss on parse error
            pass

    xlsx_url = _resolve_political_violence_resource_url(package_id)
    if not xlsx_url:
        return []

    blob = _download(xlsx_url)
    if not blob:
        return []

    try:
        cache_path.write_bytes(blob)
    except Exception:  # noqa: BLE001 — cache is best-effort
        pass

    try:
        return _parse_xlsx(country, blob)
    except Exception:  # noqa: BLE001
        return []


def _resolve_political_violence_resource_url(package_id: str) -> Optional[str]:
    """Hit ``package_show`` to find the current political-violence XLSX URL.

    The HDX URL contains the as-of-date and changes every Monday when ACLED
    publishes a new release, so we always re-discover it. The result is
    small JSON (~250 KB) so this is cheap.
    """
    try:
        url = f"{HDX_BASE}/package_show?id={package_id}"
        kwargs = {"timeout": 30, "follow_redirects": True, "verify": VERIFY_SSL}
        if PROXY:
            kwargs["proxy"] = PROXY
        with httpx.Client(**kwargs) as c:
            resp = c.get(url, headers={"User-Agent": USER_AGENT})
            if resp.status_code != 200:
                return None
            data = resp.json()
    except (httpx.HTTPError, ValueError):
        return None

    for res in (data.get("result") or {}).get("resources", []) or []:
        name = (res.get("name") or "").lower()
        fmt = (res.get("format") or "").lower()
        if fmt != "xlsx":
            continue
        if "political_violence_events_and_fatalities_by_month-year" in name:
            return res.get("url")

    # Fall back to the first XLSX resource if the political-violence one
    # isn't found (defensive — schema may change).
    for res in (data.get("result") or {}).get("resources", []) or []:
        if (res.get("format") or "").lower() == "xlsx":
            return res.get("url")
    return None


def _download(url: str) -> Optional[bytes]:
    try:
        kwargs = {"timeout": 60, "follow_redirects": True, "verify": VERIFY_SSL}
        if PROXY:
            kwargs["proxy"] = PROXY
        with httpx.Client(**kwargs) as c:
            resp = c.get(url, headers={"User-Agent": USER_AGENT})
            if resp.status_code != 200:
                return None
            return resp.content
    except httpx.HTTPError:
        return None


def _parse_xlsx(country: str, blob: bytes) -> list[_MonthlyRow]:
    """Parse the ACLED political-violence XLSX into ``_MonthlyRow``s.

    Two schemas occur in the wild on HDX:

    * **Simple** (e.g. Kenya, Bangladesh): one row per (country, month, year)
      with columns ``Country | Month | Year | Events | Fatalities``.
    * **HRP-detailed** (e.g. Sudan, Ethiopia, Somalia, Myanmar): one row per
      (country, Admin2, month, year) with columns ``Country | Admin1 |
      Admin2 | ISO3 | Admin2 Pcode | Admin1 Pcode | Month | Year | Events |
      Fatalities``. We aggregate across all Admin2 rows to recover monthly
      totals.

    Both are handled by indexing on the header row by name.

    openpyxl is imported lazily so the dependency only kicks in when the
    HDX path actually runs.
    """
    import openpyxl  # noqa: PLC0415 — lazy import

    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    sheet_name = "Data" if "Data" in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[sheet_name]

    iter_rows = ws.iter_rows(values_only=True)

    # Skip blank rows until we hit the header row that starts with "Country".
    header: Optional[tuple] = None
    for raw in iter_rows:
        if raw and any(cell not in (None, "") for cell in raw):
            if isinstance(raw[0], str) and raw[0].strip().lower() == "country":
                header = raw
                break
    if not header:
        return []

    col = {(str(h).strip().lower() if h is not None else ""): i for i, h in enumerate(header)}
    try:
        i_country = col["country"]
        i_month = col["month"]
        i_year = col["year"]
        i_events = col["events"]
        i_fatalities = col["fatalities"]
    except KeyError:
        return []

    month_lookup = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}

    # Aggregate (year, month) -> [events, fatalities] across Admin2 rows.
    bucket: dict[tuple[int, int], list[int]] = {}
    country_label = country.title()
    for raw in iter_rows:
        if not raw or i_year >= len(raw):
            continue
        try:
            c_name = (str(raw[i_country]) if raw[i_country] is not None else "").strip()
            month_str = (str(raw[i_month]) if raw[i_month] is not None else "").strip().lower()
            year_val = raw[i_year]
            year = int(year_val) if year_val is not None else 0
            events = int(raw[i_events] or 0)
            fatalities = int(raw[i_fatalities] or 0)
        except (TypeError, ValueError):
            continue
        month = month_lookup.get(month_str)
        if month is None or year < 1990:
            continue
        if c_name:
            country_label = c_name
        key = (year, month)
        if key in bucket:
            bucket[key][0] += events
            bucket[key][1] += fatalities
        else:
            bucket[key] = [events, fatalities]

    rows = [
        _MonthlyRow(
            country=country_label,
            year=year,
            month=month,
            events=ev,
            fatalities=fa,
        )
        for (year, month), (ev, fa) in bucket.items()
    ]
    rows.sort(key=lambda r: r.ordinal)
    return rows


def _aggregate_last_n_months(
    rows: list[_MonthlyRow],
    n: int,
    skip_first_n: int = 0,
) -> dict:
    """Aggregate ``rows`` over the last ``n`` months, optionally skipping
    the most recent ``skip_first_n`` months."""
    if not rows:
        return {"events": 0, "fatalities": 0}
    # rows are sorted ascending; take from the tail.
    tail = rows[-(skip_first_n + n): -skip_first_n] if skip_first_n else rows[-n:]
    return {
        "events": sum(r.events for r in tail),
        "fatalities": sum(r.fatalities for r in tail),
    }


def known_countries() -> Iterable[str]:
    """For external diagnostics — which countries the HDX path supports."""
    return COUNTRY_TO_PACKAGE.keys()
